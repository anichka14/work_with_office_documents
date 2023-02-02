"""де Проекти (Projects) – це всі проекти через підкреслення, на яких спільно
працювали дві особи, а вага (Weight) зв’язку, - кількість спільних проектів
двох осіб."""

import openpyxl
from collections import defaultdict
from itertools import combinations


def create_xlsx(filename, *sheets):
    wb = openpyxl.Workbook()  # створюємо робочу книгу
    wb.remove(wb.active)
    for title, table in sheets:
        ws = wb.create_sheet(title)  # додаємо робочий аркуш
        for row in table:
            ws.append(row)
    wb.save(filename)


def create_connections(filename):
    wb = openpyxl.load_workbook(filename)  # відкриваємо робочу книгу
    # отримаємо список робочих аркушів
    project_ws = wb["projects"]
    dct = defaultdict(list)
    for row in project_ws:
        print(row)
        dct[row[0].value].append(row[1].value)
    del dct["Project"]
    print(dct.items())

    connections_ws = wb.create_sheet("connections")
    connections_ws.append(("Person1", "Person2", "Project", "Weight"))

    for key, values in dct.items():
        for i in range(len(values)):
            for j in range(i + 1, len(values)):
                connections_ws.append((values[i], values[j], key))
                determination_of_weight(dct=dct, connections_ws=connections_ws)
    wb.save(filename)


def determination_of_weight(dct, connections_ws):
    d = defaultdict(int)
    for key, values in dct.items():
        for c in combinations(values, 2):
            # якщо кожен елемент в нашій комбінації знаходиться у values
            if all(el in values for el in c):
                d[c] += 1
    r = 2
    for row in connections_ws.iter_rows(min_row=2):
        for k, v in d.items():
            if row[0].value in k and row[1].value in k:
                connections_ws.cell(column=4, row=r, value=v)
                r += 1


if __name__ == "__main__":
    projects = (
        "projects", (
            ("Project", "Person"),
            ("project_1", "Alex"),
            ("project_1", "John"),
            ("project_1", "Kate"),
            ("project_2", "Alex"),
            ("project_2", "Stacy"),
            ("project_2", "John"),
            ("project_3", "Alex"),
            ("project_3", "Kate"),
            ("project_4", "Alex"),
            ("project_4", "Jack"),
            ("project_4", "John"),
            ("project_5", "Alex")
        )
    )
    create_xlsx("t23_11.xlsx", projects)
    create_connections("t23_11.xlsx")
