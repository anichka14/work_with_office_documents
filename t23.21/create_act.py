import openpyxl
from docxtpl import DocxTemplate


def create_invoice(datafile):
    doc = DocxTemplate("templates/template_act.docx")
    wb = openpyxl.load_workbook(datafile)  # відкриваємо файл

    # відкриваємо робочий аркуш рахунку
    ws = wb["acts"]
    act_info = {}
    for row in ws.iter_rows(min_row=2):
        act_info[row[0].value] = {"act_no": row[1].value, "act_date": row[2].value,
                                  "act_sum": row[3].value, "fields_s_id": row[4].value}

    # тепер ми маємо індентифікатор клієнта і за ним можемо шукати клієнтів
    ws = wb["fields"]
    fields_info = {}
    for row in ws.iter_rows(min_row=2):
        fields_info[row[0].value] = {"fields_name": row[1].value, "fields_address": row[2].value,
                                     "responsible_person": row[3].value, "manager": row[4].value}

    works_info = {}
    ws = wb["items"]
    for row in ws.iter_rows(min_row=2):
        works_info[row[0].value] = {"act_id": row[1].value}

    works = {}
    ws = wb["works"]
    for row in ws.iter_rows(min_row=2):
        works[row[0].value] = {"name": row[1].value}

    lst = []
    for k, v in works_info.items():
        act = act_info[works_info[k]["act_id"]]
        fields = fields_info[act["fields_s_id"]]
        works[k].update({**act, **fields})
        lst.append(works[k])
    # print(works)

    for i in range(len(lst)):
        doc.render(context=lst[i])  # заповняємо поля шаблону потрібними змінними
        doc.save(f"act{i}.docx")


if __name__ == "__main__":
    create_invoice("data.xlsx")
